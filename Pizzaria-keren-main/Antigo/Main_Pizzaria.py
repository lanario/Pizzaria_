import openpyxl
from datetime import date

def get_nome_arquivo():
    data = date.today()
    nome_arquivo = f'pizzaria {data.month}-{data.year}.xlsx'
    return nome_arquivo

def criar_planilha():
    book = openpyxl.Workbook()
    sheets = ['fornecedores', 'clientes', 'vendas']
    
    for sheet_name in sheets:
        sheet = book.create_sheet(sheet_name)
        if sheet_name == 'fornecedores':
            sheet.append(["Produto", "Quantidade", "Valor"])
        else:
            sheet.append(["Nome", "Numero", "CPF", "Endereço"])
    
    book.remove(book['Sheet'])
    return book

def carregar_planilha(nome_arquivo):
    try:
        book = openpyxl.load_workbook(nome_arquivo)
    except:
        book = criar_planilha()
    
    return book

def obter_produtos_existentes(book):
    fornecedores = book['fornecedores']
    produtos_existentes = {}
    
    for row in fornecedores.iter_rows(min_row=2, values_only=True):
        produto, quantidade, valor = row
        produtos_existentes[produto] = (quantidade, valor)
    
    return produtos_existentes

def adicionar_atualizar_produtos(produtos_existentes):
    while True:
        produto = input("Digite o nome do produto (ou digite 'sair' para sair): ")
        if produto.lower() == "sair":
            break
        
        quantidade = int(input("Digite a quantidade do produto: "))
        valor = float(input("Digite o valor do produto: "))
        
        if produto in produtos_existentes:
            quantidade_existente, valor_existente = produtos_existentes[produto]
            nova_quantidade = quantidade_existente + quantidade
            novo_valor = valor_existente + valor
            produtos_existentes[produto] = (nova_quantidade, novo_valor)
        else:
            produtos_existentes[produto] = (quantidade, valor)
    
    return produtos_existentes

def atualizar_planilha(book, produtos_existentes):
    fornecedores = book['fornecedores']
    fornecedores.delete_rows(2, fornecedores.max_row)
    
    linha_atual = 2
    for produto, dados in produtos_existentes.items():
        quantidade, valor = dados
        fornecedores.cell(row=linha_atual, column=1, value=produto)
        fornecedores.cell(row=linha_atual, column=2, value=quantidade)
        fornecedores.cell(row=linha_atual, column=3, value=valor)
        linha_atual += 1
    
    return book

def salvar_planilha(book, nome_arquivo):
    book.save(nome_arquivo)
    print("Planilha atualizada e salva com sucesso!")

def atualizar_planilha_pizzaria():
    nome_arquivo = get_nome_arquivo()
    book = carregar_planilha(nome_arquivo)
    produtos_existentes = obter_produtos_existentes(book)
    produtos_existentes = adicionar_atualizar_produtos(produtos_existentes)
    book = atualizar_planilha(book, produtos_existentes)
    salvar_planilha(book, nome_arquivo)

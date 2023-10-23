import os
import Main_Pizzaria
import openpyxl
import argparse
import requests
from tkinter import Tk, simpledialog

def carregar_clientes(planilha):
    clientes_sheet = planilha["clientes"]
    clientes_existentes = {}

    for linha in clientes_sheet.iter_rows(min_row=2, values_only=True):
        cliente = linha[0]
        numero = linha[1]
        cpf = linha[2]
        endereco = linha[3]
        clientes_existentes[cliente] = (numero, cpf, endereco)

    return clientes_existentes

def obter_endereco():
    resposta = input("O cliente sabe o CEP? S/N: ")
    if resposta.lower() == "s":
        cep = input("Digite o CEP do cliente: ")

        url = f"https://www.cepaberto.com/api/v3/cep?cep={cep}"
        try:
            headers = {'Authorization': f'Token token={obter_token_api()}'}
            response = requests.get(url, headers=headers)
            endereco = response.json()['logradouro']
            endereco += f" - {response.json()['bairro']}"
            numero = input("Digite o número do endereço do cliente: ")
            complemento = input("Digite o complemento do cliente: ")
            endereco += f", {numero} ({complemento}) - {cep}"
            print(endereco)
            return endereco
        except:
            print("CEP não encontrado.")
            pass
    else:
        return obter_endereco_sem_cep()

def obter_endereco_sem_cep():
    bairro = input("Digite o bairro do cliente: ")
    rua = input("Digite a rua do cliente: ")
    numero = input("Digite o número do endereço do cliente: ")
    complemento = input("Digite o complemento do cliente: ")
    endereco = f'{rua} - {bairro}, {numero} ({complemento})'
    return endereco

def atualizar_cliente(cliente, clientes_existentes):
    if cliente in clientes_existentes:
        resposta = input("Deseja atualizar os dados do cliente? (S/N): ")
        if resposta.lower() == "s":
            numero_existente, cpf, endereco_existente = clientes_existentes[cliente]
            novo_numero = int(input("Digite o novo número do cliente: "))
            novo_endereco = obter_endereco()
            clientes_existentes[cliente] = (novo_numero, cpf, novo_endereco)
    else:
        numero = int(input("Digite o número do cliente: "))
        cpf = input("Digite o CPF do cliente: ")
        endereco = obter_endereco()
        clientes_existentes[cliente] = (numero, cpf, endereco)

def atualizar_planilha_clientes(planilha, clientes_existentes):
    clientes_sheet = planilha["clientes"]
    clientes_sheet.delete_rows(2, clientes_sheet.max_row)

    linha_atual = 2
    for cliente, dados in clientes_existentes.items():
        numero, cpf, endereco = dados
        clientes_sheet[f"A{linha_atual}"] = cliente
        clientes_sheet[f"B{linha_atual}"] = numero
        clientes_sheet[f"C{linha_atual}"] = cpf
        clientes_sheet[f"D{linha_atual}"] = endereco
        linha_atual += 1

    planilha.save(Main_Pizzaria.get_nome_arquivo())

def carregar_arquivo_config():
    arquivo_config = os.path.join(os.path.dirname(__file__), 'config.txt')
    if os.path.exists(arquivo_config):
        with open(arquivo_config, 'r') as f:
            return f.read().strip()
    else:
        return None

def salvar_arquivo_config(token_api):
    arquivo_config = os.path.join(os.path.dirname(__file__), 'config.txt')
    with open(arquivo_config, 'w') as f:
        f.write(token_api)

def obter_token_api():
    token_api = carregar_arquivo_config()
    if not token_api:
        token_api = simpledialog.askstring("Token da API", "Digite o token da API CepAberto:")
        salvar_arquivo_config(token_api)
    return token_api

def main():
    parser = argparse.ArgumentParser(description='Gestão de Clientes da Pizzaria')
    parser.add_argument('-g', '--gui', action='store_true', help='Usar interface gráfica para o modo interativo')
    args = parser.parse_args()

    planilha = openpyxl.load_workbook(Main_Pizzaria.get_nome_arquivo())
    clientes_existentes = carregar_clientes(planilha)

    if args.gui:
        root = Tk()
        root.withdraw()

        while True:
            cliente = simpledialog.askstring("Nome do Cliente", "Digite o nome do cliente (ou digite 'sair' para sair):")
            if cliente.lower() == "sair":
                break

            atualizar_cliente(cliente, clientes_existentes)

        atualizar_planilha_clientes(planilha, clientes_existentes)
        planilha.close()
    else:
        while True:
            cliente = input("Digite o nome do cliente (ou digite 'sair' para sair): ")
            if cliente.lower() == "sair":
                break

            atualizar_cliente(cliente, clientes_existentes)

        atualizar_planilha_clientes(planilha, clientes_existentes)
        planilha.close()

if __name__ == "__main__":
    main()
